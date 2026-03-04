"""
Utilitários comuns para exportação de dados em Excel e PDF
"""
import io
import csv
from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Any, Optional

from django.http import HttpResponse

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


def formatar_valor_monetario(valor: Any) -> str:
    """Formatar valor como moeda brasileira"""
    if valor is None or valor == '':
        return 'R$ 0,00'

    try:
        if isinstance(valor, str):
            # Remove apenas R$ e espaços, mantém ponto decimal
            valor = valor.replace('R$', '').strip()
            # Se tem vírgula, assume formato brasileiro (1.568,27) e converte
            if ',' in valor:
                valor = valor.replace('.', '').replace(',', '.')

        valor_decimal = Decimal(str(valor))
        return f"R$ {valor_decimal:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except:
        return str(valor)


def exportar_excel(nome_arquivo: str, dados: List[Dict], cabecalhos: Dict[str, str] = None,
                  titulo: str = None, colunas_monetarias: List[str] = None,
                  colunas_percentuais: List[str] = None,
                  lojas_incluidas: List[str] = None) -> HttpResponse:
    """Exportar dados para Excel com formatação profissional"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl não está instalado")

    if not dados:
        raise ValueError("Nenhum dado fornecido para exportação")

    colunas_monetarias = colunas_monetarias or []
    colunas_percentuais = colunas_percentuais or []
    cabecalhos = cabecalhos or {}

    wb = Workbook()
    ws = wb.active
    ws.title = titulo or "Dados"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="65C97A", end_color="65C97A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Cabeçalhos começam na linha 1 (sem título)
    linha_atual = 1

    # Cabeçalhos
    colunas = list(dados[0].keys())
    for col_idx, coluna in enumerate(colunas, 1):
        celula = ws.cell(row=linha_atual, column=col_idx)
        celula.value = cabecalhos.get(coluna, coluna)
        celula.font = header_font
        celula.fill = header_fill
        celula.alignment = header_alignment
        celula.border = border

    linha_atual += 1

    # Dados
    for item in dados:
        for col_idx, coluna in enumerate(colunas, 1):
            celula = ws.cell(row=linha_atual, column=col_idx)
            valor = item.get(coluna, '')

            if coluna in colunas_monetarias and valor:
                try:
                    if isinstance(valor, str):
                        # Remove apenas R$ e espaços, mantém ponto decimal
                        valor_limpo = valor.replace('R$', '').strip()
                        # Se tem vírgula, assume formato brasileiro (1.568,27) e converte
                        if ',' in valor_limpo:
                            valor_limpo = valor_limpo.replace('.', '').replace(',', '.')
                    else:
                        valor_limpo = str(valor)
                    celula.value = float(valor_limpo)
                    celula.number_format = 'R$ #,##0.00'
                except:
                    celula.value = str(valor)
            elif coluna in colunas_percentuais and valor:
                try:
                    if isinstance(valor, str) and '%' in valor:
                        # Remove o símbolo % e converte para decimal
                        valor_limpo = valor.replace('%', '').strip()
                        if ',' in valor_limpo:
                            valor_limpo = valor_limpo.replace(',', '.')
                        # Converte percentual para decimal (1.30% -> 0.013)
                        celula.value = float(valor_limpo) / 100
                        celula.number_format = '0.00%'
                    elif isinstance(valor, (int, float, Decimal)):
                        # Valor já é numérico decimal (ex: -0.0622 = -6.22%)
                        celula.value = float(valor)
                        celula.number_format = '0.00%'
                    else:
                        celula.value = '' if valor is None else valor
                except:
                    celula.value = '' if valor is None else str(valor)
            else:
                celula.value = '' if valor is None else valor

            celula.border = border

        linha_atual += 1

    # Ajustar largura das colunas
    for col_idx, coluna in enumerate(colunas, 1):
        column_letter = get_column_letter(col_idx)
        max_length = len(cabecalhos.get(coluna, coluna))

        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

        # Aumentar largura mínima para evitar truncamento
        largura_calculada = max(max_length + 4, len(cabecalhos.get(coluna, coluna)) + 4)
        ws.column_dimensions[column_letter].width = min(largura_calculada, 60)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}.xlsx"'

    wb.save(response)
    return response


def remover_acentos(texto):
    """Remove acentos e caracteres especiais do texto"""
    import unicodedata
    if not texto:
        return texto

    # Normalizar e remover acentos
    texto_normalizado = unicodedata.normalize('NFD', str(texto))
    texto_sem_acentos = ''.join(c for c in texto_normalizado if unicodedata.category(c) != 'Mn')

    # Substituições específicas
    substituicoes = {
        'ç': 'c', 'Ç': 'C',
        'ñ': 'n', 'Ñ': 'N'
    }

    for original, substituto in substituicoes.items():
        texto_sem_acentos = texto_sem_acentos.replace(original, substituto)

    return texto_sem_acentos


def exportar_csv(nome_arquivo: str, dados: List[Dict], cabecalhos: Dict[str, str] = None,
                 colunas_monetarias: List[str] = None, colunas_percentuais: List[str] = None,
                 lojas_incluidas: List[str] = None) -> HttpResponse:
    """Exportar dados para CSV com formatação adequada"""
    from wallclub_core.utilitarios.log_control import registrar_log

    registrar_log('comum.utilitarios', f"DEBUG: exportar_csv chamada com nome_arquivo={nome_arquivo}")
    registrar_log('comum.utilitarios', f"DEBUG: Recebidos {len(dados) if dados else 0} itens de dados")

    if not dados:
        raise ValueError("Nenhum dado fornecido para exportação")

    cabecalhos = cabecalhos or {}
    colunas_monetarias = colunas_monetarias or []
    colunas_percentuais = colunas_percentuais or []

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}.csv"'

    response.write('\ufeff')  # BOM para Excel

    if dados:
        # Verificar se o primeiro item é um dict ou pode ser convertido
        primeiro_item = dados[0]
        registrar_log('comum.utilitarios', f"DEBUG: Tipo do primeiro item: {type(primeiro_item).__name__}")
        registrar_log('comum.utilitarios', f"DEBUG: Primeiro item: {primeiro_item}")

        if hasattr(primeiro_item, '__dict__') and not isinstance(primeiro_item, dict):
            # Se é um objeto modelo Django, converter para dict
            registrar_log('comum.utilitarios', f"DEBUG: Convertendo objeto para dict")
            primeiro_item = primeiro_item.__dict__
            dados = [item.__dict__ if hasattr(item, '__dict__') and not isinstance(item, dict) else item for item in dados]
        elif not isinstance(primeiro_item, dict):
            registrar_log('comum.utilitarios', f"ERRO: Tipo inválido - {type(primeiro_item).__name__}: {primeiro_item}", nivel='ERROR')
            raise TypeError(f"Esperado dict ou objeto com __dict__, recebido {type(primeiro_item).__name__}: {primeiro_item}")

        registrar_log('comum.utilitarios', f"DEBUG: Extraindo colunas do primeiro item")
        colunas = list(primeiro_item.keys())
        registrar_log('comum.utilitarios', f"DEBUG: Colunas encontradas: {colunas[:5]}... (total: {len(colunas)})")
        writer = csv.writer(response, delimiter=';')

        # Escrever cabeçalhos (sem acentos)
        cabecalhos_ordenados = [remover_acentos(cabecalhos.get(col, col)) for col in colunas]
        writer.writerow(cabecalhos_ordenados)

        # Escrever dados com formatação
        for item in dados:
            linha = []
            for col in colunas:
                valor = item.get(col, '')

                # Formatação especial para campos monetários
                if col in colunas_monetarias and valor:
                    try:
                        valor_float = float(valor)
                        valor = f"{valor_float:.2f}".replace('.', ',')
                    except:
                        valor = str(valor) if valor else ''
                # Formatação especial para campos percentuais
                elif col in colunas_percentuais and valor:
                    try:
                        if isinstance(valor, str) and '%' in valor:
                            valor_limpo = valor.replace('%', '').strip().replace(',', '.')
                            valor_float = float(valor_limpo)
                        else:
                            valor_float = float(valor)
                        valor = f"{valor_float:.4f}".replace('.', ',')
                    except:
                        valor = str(valor) if valor else ''
                # Formatação especial para data
                elif col == 'data_transacao' and valor:
                    try:
                        valor = valor.strftime('%d/%m/%Y') if hasattr(valor, 'strftime') else str(valor)
                    except:
                        valor = str(valor)
                else:
                    valor = '' if valor is None else (str(valor) if valor else '')

                # Remover acentos de todos os valores
                valor = remover_acentos(valor)
                linha.append(valor)

            writer.writerow(linha)

    return response


def exportar_pdf(nome_arquivo: str, dados: List[Dict], titulo: str = None,
                 colunas_monetarias: List[str] = None, colunas_percentuais: List[str] = None,
                 lojas_incluidas: List[str] = None) -> HttpResponse:
    """Exportar dados para PDF com formatação profissional"""
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab não está instalado")

    if not dados:
        raise ValueError("Nenhum dado fornecido para exportação")

    colunas_monetarias = colunas_monetarias or []
    colunas_percentuais = colunas_percentuais or []

    # Criar buffer
    buffer = io.BytesIO()

    # Configurar documento PDF em paisagem
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )

    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1,  # Centro
        textColor=colors.HexColor('#1a4480')
    )

    # Elementos do documento
    elements = []

    # Título
    if titulo:
        elements.append(Paragraph(titulo, title_style))
        elements.append(Spacer(1, 12))

    # Preparar dados da tabela
    colunas = list(dados[0].keys())

    # Cabeçalhos
    table_data = [colunas]

    # Dados
    for item in dados:
        row = []
        for coluna in colunas:
            valor = item.get(coluna, '')

            # Formatar valores monetários
            if coluna in colunas_monetarias and valor:
                try:
                    if isinstance(valor, str):
                        # Remove apenas R$ e espaços, mantém ponto decimal
                        valor_limpo = valor.replace('R$', '').strip()
                        # Se tem vírgula, assume formato brasileiro (1.568,27) e converte
                        if ',' in valor_limpo:
                            valor_limpo = valor_limpo.replace('.', '').replace(',', '.')
                    else:
                        valor_limpo = str(valor)
                    valor_float = float(valor_limpo)
                    valor = f"R$ {valor_float:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                except:
                    valor = str(valor)
            # Formatar valores percentuais
            elif coluna in colunas_percentuais and valor:
                try:
                    if isinstance(valor, str) and '%' in valor:
                        valor_limpo = valor.replace('%', '').strip().replace(',', '.')
                        valor_float = float(valor_limpo)
                    else:
                        valor_float = float(valor)
                    valor = f"{valor_float:.2f}%"
                except:
                    valor = str(valor)
            else:
                valor = '' if valor is None else str(valor)

            row.append(valor)
        table_data.append(row)

    # Calcular larguras de colunas proporcionais
    # Largura disponível na página (landscape A4 com margens)
    largura_disponivel = landscape(A4)[0] - 1*inch  # Descontar margens

    # Calcular largura proporcional para cada coluna
    num_colunas = len(colunas)
    largura_coluna = largura_disponivel / num_colunas
    col_widths = [largura_coluna] * num_colunas

    # Criar tabela com larguras definidas
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Estilo da tabela
    table.setStyle(TableStyle([
        # Cabeçalho
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a4480')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 6),

        # Dados
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 5),

        # Bordas
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        # Zebra striping
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))

    elements.append(table)

    # Rodapé com data
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}",
        styles['Normal']
    ))

    # Construir PDF
    doc.build(elements)

    # Preparar resposta
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}.pdf"'

    return response


def criar_excel_em_arquivo(caminho_arquivo: str, dados: List[Dict], cabecalhos: Dict[str, str] = None,
                          titulo: str = None, colunas_monetarias: List[str] = None,
                          colunas_percentuais: List[str] = None):
    """Criar arquivo Excel no caminho especificado"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl não está instalado")

    if not dados:
        raise ValueError("Nenhum dado fornecido para exportação")

    colunas_monetarias = colunas_monetarias or []
    colunas_percentuais = colunas_percentuais or []
    cabecalhos = cabecalhos or {}

    wb = Workbook()
    ws = wb.active
    ws.title = titulo or "Dados"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="65C97A", end_color="65C97A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Cabeçalhos começam na linha 1 (sem título)
    linha_atual = 1

    # Cabeçalhos
    colunas = list(dados[0].keys())
    for col_idx, coluna in enumerate(colunas, 1):
        celula = ws.cell(row=linha_atual, column=col_idx)
        celula.value = cabecalhos.get(coluna, coluna)
        celula.font = header_font
        celula.fill = header_fill
        celula.alignment = header_alignment
        celula.border = border

    linha_atual += 1

    # Dados
    for item in dados:
        for col_idx, coluna in enumerate(colunas, 1):
            celula = ws.cell(row=linha_atual, column=col_idx)
            valor = item.get(coluna, '')

            if coluna in colunas_monetarias and valor:
                try:
                    if isinstance(valor, str):
                        # Remove apenas R$ e espaços, mantém ponto decimal
                        valor_limpo = valor.replace('R$', '').strip()
                        # Se tem vírgula, assume formato brasileiro (1.568,27) e converte
                        if ',' in valor_limpo:
                            valor_limpo = valor_limpo.replace('.', '').replace(',', '.')
                    else:
                        valor_limpo = str(valor)
                    celula.value = float(valor_limpo)
                    celula.number_format = 'R$ #,##0.00'
                except:
                    celula.value = str(valor)
            elif coluna in colunas_percentuais and valor:
                try:
                    if isinstance(valor, str) and '%' in valor:
                        # Remove o símbolo % e converte para decimal
                        valor_limpo = valor.replace('%', '').strip()
                        if ',' in valor_limpo:
                            valor_limpo = valor_limpo.replace(',', '.')
                        # Converte percentual para decimal (1.30% -> 0.013)
                        celula.value = float(valor_limpo) / 100
                        celula.number_format = '0.00%'
                    elif isinstance(valor, (int, float, Decimal)):
                        # Valor já é numérico decimal (ex: -0.0622 = -6.22%)
                        celula.value = float(valor)
                        celula.number_format = '0.00%'
                    else:
                        celula.value = '' if valor is None else valor
                except:
                    celula.value = '' if valor is None else str(valor)
            else:
                celula.value = '' if valor is None else valor

            celula.border = border

        linha_atual += 1

    # Ajustar largura das colunas
    for col_idx, coluna in enumerate(colunas, 1):
        column_letter = get_column_letter(col_idx)
        max_length = len(cabecalhos.get(coluna, coluna))

        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

        # Aumentar largura mínima para evitar truncamento
        largura_calculada = max(max_length + 4, len(cabecalhos.get(coluna, coluna)) + 4)
        ws.column_dimensions[column_letter].width = min(largura_calculada, 60)

    # Salvar arquivo
    wb.save(caminho_arquivo)


def criar_csv_em_arquivo(caminho_arquivo: str, dados: List[Dict], cabecalhos: Dict[str, str] = None):
    """Criar arquivo CSV no caminho especificado"""
    if not dados:
        raise ValueError("Nenhum dado fornecido para exportação")

    cabecalhos = cabecalhos or {}

    with open(caminho_arquivo, 'w', newline='', encoding='utf-8') as arquivo:
        arquivo.write('\ufeff')  # BOM para Excel

        if dados:
            colunas = list(dados[0].keys())
            writer = csv.DictWriter(arquivo, fieldnames=colunas, delimiter=',')

            cabecalhos_csv = {col: cabecalhos.get(col, col) for col in colunas}
            writer.writerow(cabecalhos_csv)
            writer.writerows(dados)
